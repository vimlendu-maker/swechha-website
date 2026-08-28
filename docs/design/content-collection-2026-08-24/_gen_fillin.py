#!/usr/bin/env python3
"""The two fill-in sheets, built to match what the site actually accepts.

  1 FILL THIN PAGES  — one row per gap on a page that is already live
  2 NEW PAGES        — one row per thing that has no page, with the exact
                       fields a page is made of

Column design is not arbitrary. The build rejects a figure without
value/label/period/basis/source, and rejects a photograph with no catalogue
entry — so those become columns rather than being discovered later.
"""
import csv, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", "..", ".."))
def load(p): return json.load(open(os.path.join(ROOT, p), encoding="utf-8"))

BLANK = ""

# ════════════════════ SHEET 1 — FILL THE THIN PAGES ════════════════════
C1 = ["ID","Page","Section","Kind of answer needed",
      "What the site says now","What is missing",
      ">>> YOUR ANSWER (write here)",
      "Number: value (e.g. 700+)","Number: what it counts","Number: period (e.g. since 2010)",
      "Counted or modelled?","Who says so + date (e.g. Ashim 2026-09-02)",
      "Photo / document link","Caption for the photo (what it SHOWS)",
      "Who can answer","Priority","Status"]

r1 = []
def g(i,page,sect,kind,now,missing,who,pri):
    r1.append([i,page,sect,kind,now,missing,BLANK,BLANK,BLANK,BLANK,BLANK,BLANK,BLANK,BLANK,who,pri,"NOT STARTED"])

# ── the ten live item pages: their own declared holes ──
KINDS = ["projects","campaigns","journeys","events"]
n = 0
for kind in KINDS:
    d = os.path.join(ROOT, "data/work", kind)
    for fn in sorted(os.listdir(d)):
        j = json.load(open(os.path.join(d, fn), encoding="utf-8"))
        if not j.get("page"): continue
        for h in j.get("holes", []):
            n += 1
            what = h.get("what","")
            unlocks = h.get("unlocks","")
            kindof = ("Photograph" if any(w in what.lower() for w in
                      ("photograph","frame","picture","photo")) else
                      "Number" if any(w in unlocks.lower() for w in
                      ("number","count","year","rate","how many")) else "Text")
            g(f"A-{n:02d}", f"/work/{kind}/{j['slug']}", j["name"], kindof,
              "Page is live. " + j.get("line",""),
              what + "  ->  " + unlocks,
              "Ashim", "CRITICAL" if kindof == "Photograph" else "HIGH")

# ── the pages that are not work items ──
OTHER = [
 ("/about","The record since 2000","Text",
  "Four entries: 2000, 2004, 2016, Now.",
  "22 of 26 years are blank. Need 15-25 dated entries, each with one anchor (a report page, a photo, a partner). The nine annual reports in the repo are a year-by-year record of exactly this period.",
  "Whoever reads the reports","CRITICAL"),
 ("/about","Transparency","Document",
  "Nothing. The nine reports and the 80G certificate are on the server, linked from no page.",
  "A decision on how they appear, plus the FY each covers and one line on each. Name the five lost reports and the two gap years as gaps.",
  "Vimlendu","CRITICAL"),
 ("/about","Registration","Text",
  "Footer says 'Registered Societies Registration Act, 80G, 12A, FCRA Powered' — no numbers.",
  "Societies Act number and date; 12A number; 80G number and validity; FCRA number and validity; the exact registered legal name.",
  "Neeraj","CRITICAL"),
 ("/about","Money","Number",
  "No income or expenditure figure anywhere on the site.",
  "Three years of total income, total expenditure, and share to programmes vs admin. Already printed inside the annual reports.",
  "Neeraj","HIGH"),
 ("/about","Governance","Text",
  "Eight board members listed. Three are also staff, which the page notes and does not explain.",
  "How often the board meets, what it decides, term lengths, how members are appointed, and the policy on staff who also sit on the board.",
  "Vimlendu / Kuriakose","MEDIUM"),
 ("/about","Policies","Document",
  "None published.",
  "Child protection / safeguarding, POSH policy and committee, consent process for photographing minors. Schools ask for these before they ask about price.",
  "Ashim / Neeraj","HIGH"),
 ("/about","Awards","Text",
  "Every award is buried inside one biography.",
  "A verified list: award, awarding body, year, what it was for, link or scan. Separate Vimlendu's personal awards from Swechha's institutional ones.",
  "Vimlendu","MEDIUM"),
 ("/about","Naveen Joshua","Text",
  "Name and role only. Every other board member has a biography.",
  "100-250 words: current role, qualifications, why he is on this board. Plus a portrait if he consents.",
  "Vimlendu","MEDIUM"),
 ("/about","Board portraits","Photograph",
  "Eight staff portraits. No board portraits — the old files are 338x232 and too small.",
  "Eight portraits at 1200px or better, with each member's consent.",
  "Vimlendu","MEDIUM"),
 ("/","Impact band","Number",
  "Four figures under 'Swechha's own record'. Two of them — 6,890t 'Out of the Yamuna' and '100+ green infrastructures across 100+ schools' — are in no data file and not on /impact.",
  "Source both, or withdraw both. For 6,890t: what was weighed, over what period, by whom. For the green infrastructures: what counts as one, and over what period.",
  "Ashim / Vimlendu","CRITICAL"),
 ("/","Paper archive strip","Photograph",
  "27 year-cells, 7 real, 20 placeholders. Every year 2000-2017 is a placeholder.",
  "One datable image per year for 2000-2017, plus 2021 and 2024. Photo, poster, clipping, letter or report cover — 20 items.",
  "Office archive","CRITICAL"),
 ("/","'Do it yourself' door","Text",
  "Describes four evergreen guides — compost, balcony air-detox garden, school waste audit, self-guided river walk — and links nowhere.",
  "Either the four guides, or a decision to remove the door.",
  "Ashim / Nikhil","MEDIUM"),
 ("/farm","What runs here","Text",
  "Describes the place and five visit formats. No programme is described.",
  "The five farm training programmes: Bee Keepers Collective, Composting & Micro-enterprises, Soil Regeneration, Sustainable Agriculture Camps, Women Farmers Collective. What each trains, who attends, how many have been through, which years.",
  "Naveen / Ashim","HIGH"),
 ("/farm","Built by Mewat","Text",
  "A section titled 'Built by Mewat' that names nobody from Mewat.",
  "Two or three people who built or work the farm: name, village, what they do, one quote, a portrait. With consent.",
  "Naveen","HIGH"),
 ("/farm","Visits","Text",
  "30 school groups a year, five visit formats. No school named, no visit described.",
  "One day-visit and one overnight camp told from the school's side, with a named teacher quote and photographs. With permission.",
  "Naveen / Ashim","HIGH"),
 ("/farm","Practicalities","Text",
  "Says 90 minutes from Delhi, sleeps 100, 'not a hotel'. No costs, seasons, capacity, accessibility or safeguarding.",
  "Indicative cost basis, best months, group size limits, accessibility, safeguarding for overnight stays with minors, what a school brings.",
  "Naveen","MEDIUM"),
 ("/farm","Where it is","Text",
  "'An hour and a half from Delhi'. No map, no address.",
  "Nearest town, address or pin, how to get there. Plus a decision on whether to publish a precise location.",
  "Naveen / Vimlendu","MEDIUM"),
 ("/impact","The register","Number",
  "33 figures. Nine carry 'cumulative, no start year sourced'.",
  "Nine start years: Eco Action butterfly parks, Eco Action herb gardens, Farm School composting, Farm School honey, Monsoon Wooding cumulative trees, Gram Anubhav journeys, Gram Anubhav partners, NatureScapes journeys, She Leads Change cohort period.",
  "Ashim","HIGH"),
 ("/impact","Effect, not reach","Number",
  "All 33 figures count reach. Nothing counts effect.",
  "Any one of: tree survival rate, what changed in one school, a fellow's project outcome, per-walk demand, teachers trained.",
  "Ashim","HIGH"),
 ("/publications","The publications","Document",
  "Three items.",
  "Six more that already exist and are public: NatureScapes Manual (Sirmaur), NatureScapes Manual (Sariska), Yamuna Manual, CineGreen Manual, New Delhi Brochure, and the Green Finance material. Drive IDs are in the legacy audit.",
  "Ashim","HIGH"),
 ("/stories","Written","Text",
  "Five essays, all 2022-23. Nothing published in two years.",
  "Three or four new pieces, bylined. Best candidates from work already done: the Vasant Kunj decade, Yamunotsav, a fellow's project, a Gram Anubhav host's account. Plus: publish or drop the 2018 essay 'Learning to grow with Swechha'.",
  "Ashim / Srija / Anushka","HIGH"),
 ("/stories","Films","Document",
  "Five entries. 'Disposable' returns zero hits across all 148 indexed videos; 'Yatra' is not a film. 'Tatva' and 'Sakhi' are named in a bio and appear nowhere.",
  "A file or link for Disposable, Tatva and Sakhi — or written confirmation they are lost. Plus the Ford Foundation credit for Sakhi.",
  "Vimlendu","MEDIUM"),
 ("/stories","Videos","Document",
  "148 videos indexed in the repo; roughly 8 used on the site.",
  "A curation pass: which video belongs on which page. Several pageless items almost certainly have footage.",
  "Anushka / Ashim","MEDIUM"),
 ("/now/air, /now/yamuna, /now/heat, /now/forest-fire, /now/forest-loss, /now/climate-event","What we do about it","Document",
  "Each links Swechha's programmes. None documents an advocacy action.",
  "Any RTI, consultation response, PIL, ministry submission or official meeting on that issue: date, addressee, what was asked, what happened.",
  "Vimlendu / Kuriakose","HIGH"),
 ("/work/projects/influence","The fellows","Text",
  "10 fellowships a year since 2010. No fellow named.",
  "For 5-10 fellows: name, city/state, fellowship year, project run, one outcome, a portrait, a 50-word update, written permission. Plus how long a fellowship runs.",
  "Ashim","CRITICAL"),
 ("Site-wide","Voices","Text",
  "The only quoted voice on 35 pages is the Executive Director, twice, on /farm.",
  "8-12 short attributed quotes: 2 students, 2 teachers, 2 fellows, 2 village hosts, 1 volunteer, 1 funder. Exact words, name, role, place, year, written consent.",
  "Programme team","HIGH"),
 ("Site-wide","Media coverage","Document",
  "No record that Swechha has ever been covered by anyone.",
  "Scans and links: CNN 'Be the Change' (2007-08), India Today and Outlook top-25 (2004), International Youth Foundation (2004), UN General Assembly (2011), NDTV/BBC/CBC broadcasts. The old site's news archive was 59 empty shells, so this must come from paper.",
  "Vimlendu","CRITICAL"),
 ("Site-wide","Partners and funders","Text",
  "20 organisations named, all inside five item pages. No logos, no years, no permission record.",
  "Per organisation: relationship type, project, years, one line, logo file, website, and written permission to display. Start from the funder lists inside the nine annual reports.",
  "Neeraj / Vimlendu","HIGH"),
 ("Site-wide","Schools","Text",
  "250+ claimed. Five named.",
  "20-30 named partner schools, not 250: name, city, state, programme, years, participants, and permission.",
  "Ashim","CRITICAL"),
 ("/explore","Whole page","Decision",
  "Live, crawlable, serving the OLD design with the old navigation and three 'nothing published yet' empty states.",
  "A decision: redirect it to /stories, or delete it. No content needed.",
  "Vimlendu","CRITICAL"),
]
for i,(page,sect,kind,now,miss,who,pri) in enumerate(OTHER, start=1):
    g(f"B-{i:02d}", page, sect, kind, now, miss, who, pri)

# ════════════════════ SHEET 2 — NEW PAGES ════════════════════
C2 = ["ID","Item","Kind","On the site today","Can it have a page?",
      "MINIMUM to earn a page",
      "Name to use","One line (about 110 characters)","Intro — two sentences",
      "Year started","Year ended, or 'ongoing'","Where it happens / happened",
      "What it sets out to do (3-4 short points)","How it works (2-3 points)",
      "Activities (2-4, a sentence each)","Who it is for (3-4 points)",
      "What it has done (3-4 points)",
      "Figures — value | what it counts | period | counted or modelled | who said it + date",
      "Schools involved","Partners","Funders",
      "Photo links (Drive)","Captions for those photos",
      "Statement line (under 34 characters, NO digits)",
      "What we still cannot say","Who can answer","Status"]

r2 = []
def n2(i,item,kind,today,canpage,minimum,who,prefill=None):
    row=[i,item,kind,today,canpage,minimum]+[BLANK]*(len(C2)-8)+[who,"NOT STARTED"]
    if prefill:
        for col,val in prefill.items(): row[C2.index(col)] = val
    r2.append(row)

PROMOTE = "YES — but it is a RULING. Supply the minimum below and the page is switched on in one line."
NEVER   = "NO — events never get a detail page (closed ruling). This becomes a RICHER ROW instead."
NEW     = "YES — new item. Needs a file, then the same ruling."

# 2a. Ruled rows that can be promoted
for i,(item,kind,today,minimum,who) in enumerate([
 ("We for Yamuna","campaigns","Card on /work/campaigns. Founding campaign, 2000.",
  "One dated action and one stated demand — anything that is not the year 2000.","Vimlendu / Ashim"),
 ("Delhi I Can't See You","campaigns","Card. Cross-linked from /now/air as our air campaign.",
  "When it started, what it asks for, one thing it has done, and one photograph from a real action.","Ashim"),
 ("This Girl Can","campaigns","Card.",
  "When it started, what it asked for, one thing it did — AND whether it is its own campaign or the name She Leads Change / ME to WE campaign under.","Ashim"),
 ("No more Waste Hills","campaigns","Card.",
  "Which landfill, what was asked and of whom, what year. And whether the CityScapes landfill walk is part of it.","Ashim"),
 ("No Plastic","campaigns","Card.",
  "Whether it is a ban, a substitution or a shop. Plus when it ran and what changed. A refusal counts as an outcome.","Ashim"),
 ("Park Restoration","campaigns","Card.",
  "First: is this the same work as Eco Action under another name? If yes it becomes a section there. If no: one park, one year, one demand.","Ashim"),
 ("Sustainable Shopping","campaigns","Card.",
  "Who it spoke to and what it asked them to buy instead; the years; and one line on whether it is connected to Green the Map.","Ashim"),
 ("She Leads Change","projects","Card with three sourced figures — closest to page-ready.",
  "One sentence on how it differs from ME to WE (or a merge), a current cohort size, and the citation for the 2018 ELC Bright Promise Award.","Ashim"),
 ("Food systems, with UNEP","projects","Card. Our own sentence is in the future tense under a heading reading 'What is running'.",
  "Has it started? A start date, a school count, one word for UNEP's role (funder / technical partner / co-designer), and one dated deliverable.","Ashim"),
], start=1):
    n2(f"P-{i:02d}", item, kind, today, PROMOTE, minimum, who)

# 2b. Events — richer rows only
for i,(item,today,minimum,who) in enumerate([
 ("Yamunotsav","Row. Nine editions 2006-2014, no photograph.",
  "Access to the Yamunotsav Drive folder; what happened at one; roughly how many came; why 2014 was the last, and whether it returns.","Vimlendu"),
 ("Yamuna Shramdaan","Row. We cannot say when one last ran.",
  "Whether they still run, roughly how many have run, which bank stretches, and the archive folder. This unblocks the volunteer ask on /act.","Ashim"),
 ("Cyclothon","Row, 43 words of data.",
  "One line on what happens, roughly how many editions, which years, and the archive folder.","Ashim"),
 ("Greenathon","Row, 42 words of data.",
  "One line on what happens, roughly how many editions, which years.","Ashim"),
], start=1):
    n2(f"E-{i:02d}", item, "events", today, NEVER, minimum, who)

# 2c. Brand new items
for i,(item,kind,minimum,who) in enumerate([
 ("The Remakery","projects","What it is, where, when it opened, the weekly 'One Night Stand' events, what it makes, who runs it, and whether it is open now. /remakery currently 404s.","Nikhil / Ashim"),
 ("Teacher Training","projects","Verified number of teachers trained, over what period, in which schools, what the training covers, and whether it runs standalone or only inside Bridge the Gap.","Ashim"),
 ("Green Finance (with IGES)","projects","Study period, method, what the 93 social enterprises were, the findings, the published output, and how it relates to the IGES report already on /publications.","Vimlendu"),
 ("Circular Economy / Marine Litter & EPR","projects","Project period, GIZ's role, MoEFCC's role, the deliverables, and clearance to use both names and the 10 infographics. Eight GIZ images already sit unused in the repo.","Vimlendu / Ashim"),
 ("Women & Non-Traditional Livelihoods","projects","Programme years, the British Council YWSEDP relationship, what Udaan / MOM candles / Lunchbox 17 were, the 25,000-meals figure and its source, and what happened when it became Million Kitchen.","Ashim"),
 ("Pagdandi","projects","The 2007 start, the open-air school at Kudsia Ghat, the Jagdamba Camp school, the Kitaab Ghar library, and the RTE campaign with its 150+ children figure. This is ME to WE's first chapter.","Vimlendu / Ashim"),
 ("Films & Documentaries","projects","Broadcast history with dates and channels (CNN International, NDTV, BBC, CBC), the Ford Foundation relationship for Sakhi, and director/producer credits.","Vimlendu"),
 ("Bee Keepers Collective Training","projects","What it trains, who attends, how many have been through, which years, any partner. Note the old site said 500 kg and 600 litres in adjacent paragraphs — neither is usable without a source.","Naveen"),
 ("Composting & Micro-enterprises","projects","What it trains, who attends, how many, which years.","Naveen"),
 ("Soil Regeneration Training","projects","What it trains, who attends, how many, which years. The old counter (9,385) is contaminated placeholder data — do not reuse it.","Naveen"),
 ("Sustainable Agriculture Training Camps","projects","What it trains, who attends, how many, which years.","Naveen"),
 ("Women Farmers Collective","projects","What it trains, who attends, how many, which years.","Naveen"),
 ("Learning Communities","projects","What it is, which communities, which years, and how it relates to She Leads Change's published '300 girls in the wider Learning Communities cohort'.","Ashim"),
 ("Green Creeps","projects","Does it still exist? If yes: what it is, years, scale. If no: one timeline line instead of a page.","Ashim"),
 ("Road to Leadership","projects","Does it still exist? If yes: what it is, years, scale. If no: one timeline line instead of a page.","Ashim"),
 ("Brake Even","campaigns","What it was, which years, what it asked for. Appears on zero pages today.","Ashim"),
 ("Micro Grants / US Embassy Micro Grants","projects","What was granted, to whom, how many, which years, and the US Embassy relationship.","Ashim"),
 ("Green Exposures / Eco Walks","journeys","Is this the ancestor of CityScapes? If yes, fold it in. If separate: what it is, years, scale.","Ashim"),
 ("Future Lifestyles Project","projects","What it was and how it relates to the IGES New Delhi scenario already on /publications.","Vimlendu"),
 ("Air Pollution Campaigns","campaigns","Is this the ancestor of Delhi I Can't See You? Fold, or describe separately.","Ashim"),
 ("Cycles for Sustainability","events","Is this Cyclothon under another name? Fold, or describe separately.","Ashim"),
 ("Dairy Cooperative and Cow Rearing","projects","Listed on the old site's Programs page but never had a page even there. Does it exist as a programme? /farm names 20 cows.","Naveen"),
 ("US Dept of State / podcast & masterclass series","projects","What the programme was, the US Dept of State relationship, which years. The podcasts are already on /stories; the programme behind them is not described.","Vimlendu / Anushka"),
], start=1):
    n2(f"N-{i:02d}", item, kind, "Nothing. Old URL redirects to a parent index.", NEW, minimum, who)

# ════════════════════ write ════════════════════
def csvout(name, cols, rows):
    with open(os.path.join(HERE, name), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL); w.writerow(cols); w.writerows(rows)
    print(name, len(rows), "rows")

csvout("FILL-1-thin-pages.csv", C1, r1)
csvout("FILL-2-new-pages.csv", C2, r2)

# ── workbook ──
MUTED="FF615B50"; INK="FF141310"
HEAD=PatternFill("solid",fgColor="FFECEBE8")
ANSWER=PatternFill("solid",fgColor="FFFDF6E3")
CRIT=PatternFill("solid",fgColor="FFFBE9EC")
NEVERF=PatternFill("solid",fgColor="FFF0EFEC")
THIN=Side(style="thin",color="FFDEDDD9")

wb=Workbook(); ws=wb.active; ws.title="READ FIRST"
guide=[("Swechha — two jobs, two sheets","h1"),("",""),
 ("Sheet 1 fills in pages that are already live but thin. Sheet 2 creates pages for work that is not on the site at all.","p"),("",""),
 ("TWO HARD RULES FROM THE WEBSITE ITSELF","h2"),
 ("1. EVERY NUMBER NEEDS FOUR THINGS or the site refuses to publish it: the value, what it counts, the period it covers, and whether it was COUNTED or is an ESTIMATE. Plus who said it and on what date. That is why those are separate columns — please do not merge them into a sentence.","p"),
 ("2. EVENTS NEVER GET THEIR OWN PAGE. That is a closed decision on the site, not an oversight. Yamunotsav, Cyclothon, Greenathon and Yamuna Shramdaan can become much richer entries with photographs and dates, but they stay as entries on the Events page. Everything you supply for them is still used.","p"),("",""),
 ("HOW A NEW PAGE ACTUALLY HAPPENS","h2"),
 ("Whether something is a page or an entry is a decision, not a data field — the site is built to refuse a thin page on the grounds that a rich entry is better. So each row in Sheet 2 has a 'MINIMUM to earn a page' column. Fill that much and the page gets switched on. There is precedent: Monsoon Wooding became a page the moment one question about how tree survival is counted was answered.","p"),("",""),
 ("WHAT TO DO","h2"),
 ("Work along the row. The wide cream column in Sheet 1 is where you write. In Sheet 2 each column is literally a part of the finished page, so what you write is close to what a reader will see.","p"),
 ("Photographs: paste a Drive link, never the image itself. Every photograph also needs a caption saying what it SHOWS — not what it stands for. The site will not publish a picture without one.","p"),
 ("If you cannot verify something, write CANNOT VERIFY and move on. Do not estimate. The site's whole argument is that its numbers can be checked, and one invented figure costs more than twenty missing ones.","p"),
 ("Do not copy numbers off the old website. Four of those old pages were unfinished templates sharing one set of invented numbers, and they leaked onto two real pages. The writing on them is good — the counters are not.","p"),("",""),
 ("SEND IT BACK IN BATCHES","h2"),
 ("Ten or twenty finished rows at a time, rather than waiting for the whole sheet. Each batch gets transcribed into the site and the pages rebuilt, so you will see it working while the rest is still being collected.","p"),
]
for i,(t,k) in enumerate(guide,start=1):
    c=ws.cell(row=i,column=1,value=t)
    c.font=Font(name="Calibri",size=17 if k=="h1" else 11,bold=k in("h1","h2"),
                color=MUTED if k=="h2" else INK)
    c.alignment=Alignment(wrap_text=True,vertical="top")
ws.column_dimensions["A"].width=124
ws.sheet_view.showGridLines=False

for title,cols,rows in [("1 FILL THIN PAGES",C1,r1),("2 NEW PAGES",C2,r2)]:
    s=wb.create_sheet(title)
    s.append(cols)
    for r in rows: s.append(r)
    for ci,name in enumerate(cols,start=1):
        h=s.cell(row=1,column=ci); h.font=Font(name="Calibri",size=10,bold=True,color=MUTED)
        h.fill=HEAD; h.alignment=Alignment(wrap_text=True,vertical="bottom")
        n=name.lower()
        s.column_dimensions[get_column_letter(ci)].width = (
            8 if n=="id" else 14 if n in("kind","status","priority") else
            60 if (">>>" in name or "minimum" in n or "what is missing" in n or
                   "figures" in n or "what the site says now" in n or
                   "activities" in n or "what it sets out" in n) else
            22 if ("number:" in n or "counted" in n or "who say" in n or "year" in n) else 34)
    s.row_dimensions[1].height=58
    s.freeze_panes="C2" if title.startswith("1") else "C2"
    s.auto_filter.ref=f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
    for ri in range(2,len(rows)+2):
        for ci in range(1,len(cols)+1):
            c=s.cell(row=ri,column=ci)
            c.alignment=Alignment(wrap_text=True,vertical="top")
            c.font=Font(name="Calibri",size=10); c.border=Border(bottom=THIN)
    last=len(rows)+1
    for key,opts in (("Priority",["CRITICAL","HIGH","MEDIUM","LOW"]),
                     ("Status",["NOT STARTED","IN PROGRESS","DONE","CANNOT VERIFY","NOT FOR PUBLICATION"]),
                     ("Counted or modelled?",["counted","modelled"])):
        if key in cols:
            col=get_column_letter(cols.index(key)+1)
            dv=DataValidation(type="list",formula1='"'+",".join(opts)+'"',allow_blank=True)
            s.add_data_validation(dv); dv.add(f"{col}2:{col}{last}")
    if ">>> YOUR ANSWER (write here)" in cols:
        ac=cols.index(">>> YOUR ANSWER (write here)")+1
        for ri in range(2,last+1): s.cell(row=ri,column=ac).fill=ANSWER
    if "Priority" in cols:
        pc=cols.index("Priority")+1
        for ri in range(2,last+1):
            if s.cell(row=ri,column=pc).value=="CRITICAL": s.cell(row=ri,column=pc).fill=CRIT
    if "Can it have a page?" in cols:
        cc=cols.index("Can it have a page?")+1
        for ri in range(2,last+1):
            if str(s.cell(row=ri,column=cc).value).startswith("NO"):
                s.cell(row=ri,column=cc).fill=NEVERF

wb.save(os.path.join(HERE,"SWECHHA-fill-in.xlsx"))
print("SWECHHA-fill-in.xlsx — 3 tabs")
