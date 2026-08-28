#!/usr/bin/env python3
"""Builds the two single-file deliverables from the seven CSVs.

  00-ALL-SHEETS-MERGED.csv        one flat CSV, union schema, `Sheet` discriminator
  swechha-content-collection.xlsx one workbook, 9 tabs, dropdowns, filters

Also adds two working columns to the master sheet — Wave and Assigned To — so
the sheet is a work plan that three people can split, not just a list.
"""
import csv, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

FILES = [
    ("01", "01-master-content-collection.csv", "Master sheet"),
    ("02", "02-project-master-list.csv", "Projects"),
    ("03", "03-partner-master-list.csv", "Partners"),
    ("04", "04-school-institution-master-list.csv", "Schools"),
    ("05", "05-impact-numbers-master-list.csv", "Impact numbers"),
    ("06", "06-photo-visual-asset-list.csv", "Photos"),
    ("07", "07-publication-media-archive.csv", "Publications & media"),
    ("08", "08-photos-we-already-have.csv", "Photos we HAVE"),
]

PRIORITY = ["CRITICAL", "HIGH", "MEDIUM", "LOW"]
STATUS = ["NOT STARTED", "SEARCHING", "AVAILABLE", "NEEDS VERIFICATION",
          "READY FOR WEBSITE", "NOT AVAILABLE", "NOT FOR PUBLICATION"]

# ─────────────────────────────────────────────────────────────────────────────
# WAVES. The only judgement in this file, and it is deliberate: a wave says
# WHAT KIND OF WORK a row is, which is what decides who can do it.
#   1 — the material is already in this repository or already public. A reader,
#       not a researcher. No fieldwork, no permission, no travel.
#   2 — one person knows the answer. A conversation or an email closes it.
#   3 — an archive dig, a photo shoot, a consent, or an owner ruling.
# ─────────────────────────────────────────────────────────────────────────────
WAVE_1 = {"T-01","T-02","T-04","H-01","P-02","K-02","K-06","N-04","X-01","X-02","X-03","X-04",
          "K-01","K-03","M-10"}
WAVE_3 = {"T-05","T-07","T-08","H-02","H-03","H-04","H-06",
          "W-08","W-11",
          "D-04","D-06","D-08","D-09",
          "P-01","P-03","P-04","P-05",
          "F-01","F-02","F-03","F-04","F-05","F-06","F-07","F-08","F-09","F-10","F-11","F-12",
          "J-01","J-02","J-03","J-04",
          "R-02","R-03","R-05","X-05",
          "M-01","M-04","M-05","M-06","M-07"}

def wave_for(rid, status, internally):
    if rid in WAVE_1: return "1 — already in repo / already public"
    if rid in WAVE_3: return "3 — archive, photography, consent or owner ruling"
    if status == "AVAILABLE" or internally.startswith("YES — already in repo"):
        return "1 — already in repo / already public"
    return "2 — one conversation closes it"

# ── augment the master sheet ─────────────────────────────────────────────────
mp = os.path.join(HERE, FILES[0][1])
with open(mp, encoding="utf-8") as f:
    master = list(csv.reader(f))
head, body = master[0], master[1:]
if "Wave" not in head:
    i_status = head.index("Status")
    i_int = head.index("Existing Internally?")
    head = head[:1] + ["Wave", "Assigned To"] + head[1:]
    body = [r[:1] + [wave_for(r[0], r[i_status], r[i_int]), ""] + r[1:] for r in body]
    with open(mp, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL)
        w.writerow(head)
        w.writerows(body)
    print("master sheet: added Wave + Assigned To")

# ── load everything ──────────────────────────────────────────────────────────
sheets = []
for num, fn, tab in FILES:
    with open(os.path.join(HERE, fn), encoding="utf-8") as f:
        rows = list(csv.reader(f))
    sheets.append((num, tab, rows[0], rows[1:]))

# ── 00. one flat CSV ─────────────────────────────────────────────────────────
union, seen = [], set()
for _, _, cols, _ in sheets:
    for c in cols:
        if c not in seen:
            seen.add(c); union.append(c)
MERGED = ["Sheet", "Row ID"] + union

def row_id(num, cols, r, idx):
    if "ID" in cols: return r[cols.index("ID")]
    for k in ("Project Name", "Organisation", "Institution Name", "Metric", "Asset Needed", "Title"):
        if k in cols: return f"{num}-{idx+1:03d}"
    return f"{num}-{idx+1:03d}"

out = []
for num, tab, cols, rows in sheets:
    for i, r in enumerate(rows):
        d = dict(zip(cols, r))
        out.append([f"{num} · {tab}", row_id(num, cols, r, i)] + [d.get(c, "") for c in union])

with open(os.path.join(HERE, "00-ALL-SHEETS-MERGED.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f, quoting=csv.QUOTE_ALL)
    w.writerow(MERGED)
    w.writerows(out)
print("00-ALL-SHEETS-MERGED.csv", len(out), "rows,", len(MERGED), "columns")

# ── the workbook ─────────────────────────────────────────────────────────────
INK = "FF141310"; MUTED = "FF615B50"
HEAD_FILL = PatternFill("solid", fgColor="FFECEBE8")
CRIT_FILL = PatternFill("solid", fgColor="FFFBE9EC")
HIGH_FILL = PatternFill("solid", fgColor="FFFCF3DF")
WAVE1_FILL = PatternFill("solid", fgColor="FFE8F3EC")
THIN = Side(style="thin", color="FFDEDDD9")
BORDER = Border(bottom=THIN)

wb = Workbook()

# START HERE tab
ws = wb.active
ws.title = "START HERE"
guide = [
    ("Swechha content collection — 24 August 2026", "h1"),
    ("", ""),
    ("Eight sheets, 447 rows. Every row is one specific thing to find, not a topic. Sheets 02, 03, 05, 07 and 08 are GENERATED from the website itself, so the 'what we already have' columns cannot drift from what swechha.in actually says today.", "p"),
    ("", ""),
    ("HOW TO USE IT", "h2"),
    ("1. Filter the Master sheet by Wave. Wave 1 needs no fieldwork — the material is already in the repository or already public. Start there.", "p"),
    ("2. Put a name in 'Assigned To'. Three people can run this in parallel because the waves need three different kinds of person.", "p"),
    ("3. Update 'Status' from the dropdown as you go. Leave a row at NOT AVAILABLE rather than deleting it — a named gap is more useful than a tidy list.", "p"),
    ("Before filling sheet 06, check tab 08 — it lists every photograph already on disk, with its alt text and whether any page uses it. Eleven usable frames are currently used nowhere. Do not send anyone out to photograph something we already have.", "p"),
    ("", ""),
    ("4. Paste links, not files. For photographs and documents, put a Drive or Dropbox link in the last column. Do not embed images in the spreadsheet.", "p"),
    ("", ""),
    ("THE THREE RULES", "h2"),
    ("A named gap beats a tidy list. The website already admits what it does not have. Do not close a hole by deleting the admission.", "p"),
    ("Port prose, never legacy counters. Four old project pages were unfinished templates sharing one counter set (2740/4751/1260/9385) and it contaminated two real pages. Every old number needs a source before it goes back on the site.", "p"),
    ("Twenty named beats two hundred and fifty counted. For schools, fellows, host organisations and partners: a small verified list with permission is worth more than a large claim.", "p"),
    ("", ""),
    ("DO NOT INVENT ANYTHING", "h2"),
    ("If a number, a partner, a year or an outcome cannot be verified, write TO BE VERIFIED in the cell. An unverifiable figure on the website is worse than a gap, because the site's whole argument is that its numbers can be checked.", "p"),
    ("", ""),
    ("THE SHEETS", "h2"),
]
for num, tab, cols, rows in sheets:
    guide.append((f"{num} · {tab} — {len(rows)} rows", "p"))
guide += [
    ("", ""),
    ("WHAT HAPPENS TO A FILLED SHEET", "h2"),
    ("The website's content lives in JSON files, not in this spreadsheet. A filled sheet is transcribed into those files, the page generators are re-run, and the result is reviewed and deployed. That step is not automatic and should not be: the build refuses a figure with no source and a photograph with no catalogue entry, which is the protection that keeps the site trustworthy.", "p"),
]
for i, (text, kind) in enumerate(guide, start=1):
    c = ws.cell(row=i, column=1, value=text)
    if kind == "h1":
        c.font = Font(name="Calibri", size=17, bold=True, color=INK)
    elif kind == "h2":
        c.font = Font(name="Calibri", size=11, bold=True, color=MUTED)
    else:
        c.font = Font(name="Calibri", size=11, color=INK)
    c.alignment = Alignment(wrap_text=True, vertical="top")
ws.column_dimensions["A"].width = 128
ws.sheet_view.showGridLines = False

# data tabs
for num, tab, cols, rows in sheets:
    ws = wb.create_sheet(f"{num} {tab}"[:31])
    ws.append(cols)
    for r in rows:
        ws.append(r)

    for c in range(1, len(cols) + 1):
        h = ws.cell(row=1, column=c)
        h.font = Font(name="Calibri", size=10, bold=True, color=MUTED)
        h.fill = HEAD_FILL
        h.alignment = Alignment(wrap_text=True, vertical="bottom")
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"

    # widths: narrow for keys, wide for prose
    for ci, name in enumerate(cols, start=1):
        n = name.lower()
        if n in ("id", "wave", "priority", "status"): wdt = 17
        elif "assigned" in n: wdt = 16
        elif any(k in n for k in ("year", "permission", "available", "orientation", "type", "kind")): wdt = 22
        elif any(k in n for k in ("gap", "needed", "details", "missing", "notes", "description",
                                  "recommended", "specific", "nature", "key numbers", "activity",
                                  "alt text", "short")): wdt = 54
        else: wdt = 30
        ws.column_dimensions[get_column_letter(ci)].width = wdt

    for ri in range(2, len(rows) + 2):
        for ci in range(1, len(cols) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name="Calibri", size=10)
            cell.border = BORDER

    # dropdowns + tinting
    last = len(rows) + 1
    for key, opts in (("Priority", PRIORITY), ("Status", STATUS)):
        if key in cols:
            col = get_column_letter(cols.index(key) + 1)
            dv = DataValidation(type="list", formula1='"' + ",".join(opts) + '"', allow_blank=True)
            dv.error = f"Use one of: {', '.join(opts)}"
            ws.add_data_validation(dv)
            dv.add(f"{col}2:{col}{last}")

    if "Priority" in cols:
        pc = cols.index("Priority") + 1
        for ri in range(2, last + 1):
            v = ws.cell(row=ri, column=pc).value
            fill = CRIT_FILL if v == "CRITICAL" else HIGH_FILL if v == "HIGH" else None
            if fill: ws.cell(row=ri, column=pc).fill = fill
    if "Wave" in cols:
        wc = cols.index("Wave") + 1
        for ri in range(2, last + 1):
            if str(ws.cell(row=ri, column=wc).value or "").startswith("1"):
                ws.cell(row=ri, column=wc).fill = WAVE1_FILL

path = os.path.join(HERE, "swechha-content-collection.xlsx")
wb.save(path)
print("swechha-content-collection.xlsx", len(sheets) + 1, "tabs")
